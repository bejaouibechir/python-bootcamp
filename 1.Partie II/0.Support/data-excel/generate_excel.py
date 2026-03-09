"""
But :
Créer en une seule exécution tous les fichiers Excel échantillons (inputs)
utilisés par les exercices Excel (questions 1 à 20).

Ce que fait ce script :
- Crée un dossier data s’il n’existe pas
- Génère tous les fichiers Excel nécessaires (q01 à q20)
- Affiche un récapitulatif des fichiers créés

Prérequis :
- Installer openpyxl

Commandes à exécuter :
- pip install openpyxl
- python generate_all_excel_inputs.py

Résultat :
- Tous les fichiers Excel sont créés dans le dossier data
"""

from pathlib import Path
from datetime import datetime
from openpyxl import Workbook


def ensure_data_dir() -> Path:
    # Crée le dossier data si nécessaire
    out_dir = Path("data")
    out_dir.mkdir(exist_ok=True)
    return out_dir


def save_wb(wb: Workbook, path: Path, created: list[Path]) -> None:
    # Enregistre le classeur et mémorise le fichier généré
    wb.save(path)
    created.append(path)


def q01_ids(out_dir: Path, created: list[Path]) -> None:
    # q01 : identifiants numériques simples
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "policyID"
    ws["A2"] = 119736
    ws["A3"] = 448094

    save_wb(wb, out_dir / "q01_ids.xlsx", created)


def q02_maxrow(out_dir: Path, created: list[Path]) -> None:
    # q02 : dernière ligne “touchée” puis vidée (max_row peut rester élevé)
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "col1"
    ws["A2"] = "ok"

    ws["A100"] = "temp"
    ws["A100"] = None  # la cellule est vidée, mais la trace peut rester

    save_wb(wb, out_dir / "q02_maxrow.xlsx", created)


def q03_big(out_dir: Path, created: list[Path]) -> None:
    # q03 : fichier plus volumineux pour mesurer le coût de lecture
    wb = Workbook()
    ws = wb.active

    # En-têtes (10 colonnes)
    for c in range(1, 11):
        ws.cell(row=1, column=c).value = f"h{c}"

    # 5000 lignes de données
    for r in range(2, 5002):
        for c in range(1, 11):
            ws.cell(row=r, column=c).value = r * 100 + c

    save_wb(wb, out_dir / "q03_big.xlsx", created)


def q04_readonly(out_dir: Path, created: list[Path]) -> None:
    # q04 : volumétrie pour comparer read_only=True vs lecture normale
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "statecode", "amount"])

    for i in range(1, 30001):
        ws.append([i, "FL" if i % 2 == 0 else "TX", i * 0.5])

    save_wb(wb, out_dir / "q04_readonly.xlsx", created)


def q05_dtype(out_dir: Path, created: list[Path]) -> None:
    # q05 : une valeur manquante dans policyID peut influencer l’inférence de type
    wb = Workbook()
    ws = wb.active
    ws.append(["policyID", "statecode"])
    ws.append([119736, "FL"])
    ws.append([None, "FL"])
    ws.append([448094, "TX"])

    save_wb(wb, out_dir / "q05_dtype.xlsx", created)


def q06_usecols(out_dir: Path, created: list[Path]) -> None:
    # q06 : 20 colonnes, on pourra tester usecols pour ne lire que 2 colonnes
    wb = Workbook()
    ws = wb.active

    headers = [f"c{i}" for i in range(1, 21)]
    ws.append(headers)

    for r in range(1, 10001):
        ws.append([r * i for i in range(1, 21)])

    save_wb(wb, out_dir / "q06_usecols.xlsx", created)


def q07_nrows(out_dir: Path, created: list[Path]) -> None:
    # q07 : 50 000 lignes, utile pour tester nrows
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "value"])

    for i in range(1, 50001):
        ws.append([i, f"v{i}"])

    save_wb(wb, out_dir / "q07_nrows.xlsx", created)


def q08_dups(out_dir: Path, created: list[Path]) -> None:
    # q08 : doublons sur policyID
    wb = Workbook()
    ws = wb.active
    ws.append(["policyID", "statecode"])
    ws.append([100, "FL"])
    ws.append([101, "TX"])
    ws.append([100, "FL"])
    ws.append([102, "TX"])
    ws.append([101, "TX"])

    save_wb(wb, out_dir / "q08_dups.xlsx", created)


def q09_emptyrows(out_dir: Path, created: list[Path]) -> None:
    # q09 : ligne totalement vide et ligne partiellement vide
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b", "c"])
    ws.append([1, 2, 3])
    ws.append([None, None, None])
    ws.append([4, None, 6])

    save_wb(wb, out_dir / "q09_emptyrows.xlsx", created)


def q10_schema(out_dir: Path, created: list[Path]) -> None:
    # q10 : schéma incomplet, colonne county absente volontairement
    wb = Workbook()
    ws = wb.active
    ws.append(["policyID", "statecode"])
    ws.append([1, "FL"])

    save_wb(wb, out_dir / "q10_schema.xlsx", created)


def q11_dirty_headers(out_dir: Path, created: list[Path]) -> None:
    # q11 : en-têtes sales (espaces, casse, doublon logique)
    wb = Workbook()
    ws = wb.active
    ws.append([" PolicyID ", "STATECODE", "statecode"])
    ws.append([1, "FL", "FL"])

    save_wb(wb, out_dir / "q11_dirty_headers.xlsx", created)


def q12_dates(out_dir: Path, created: list[Path]) -> None:
    # q12 : dates sous formes différentes (datetime et texte)
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "date_ok", "date_text"])
    ws.append([1, datetime(2026, 3, 1), "2026-03-01"])
    ws.append([2, datetime(2026, 3, 2), "03/02/2026"])

    save_wb(wb, out_dir / "q12_dates.xlsx", created)


def q13_comma(out_dir: Path, created: list[Path]) -> None:
    # q13 : montants texte avec virgule (format fréquent)
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "amount_text"])
    ws.append([1, "12,5"])
    ws.append([2, "100,00"])
    ws.append([3, "7,25"])

    save_wb(wb, out_dir / "q13_comma.xlsx", created)


def q14_sheets(out_dir: Path, created: list[Path]) -> None:
    # q14 : classeur multi-feuilles (clients + orders)
    wb = Workbook()

    clients = wb.active
    clients.title = "clients"
    clients.append(["client_id", "name"])
    clients.append([1, "A"])
    clients.append([2, "B"])

    orders = wb.create_sheet("orders")
    orders.append(["order_id", "client_id", "amount"])
    orders.append([10, 1, 50])
    orders.append([11, 2, 70])

    save_wb(wb, out_dir / "q14_sheets.xlsx", created)


def q16_orphans(out_dir: Path, created: list[Path]) -> None:
    # q16 : commandes orphelines (client_id inexistant)
    wb = Workbook()

    clients = wb.active
    clients.title = "clients"
    clients.append(["client_id", "name"])
    clients.append([1, "A"])
    clients.append([2, "B"])

    orders = wb.create_sheet("orders")
    orders.append(["order_id", "client_id", "amount"])
    orders.append([10, 1, 50])
    orders.append([11, 99, 70])  # orpheline

    save_wb(wb, out_dir / "q16_orphans.xlsx", created)


def q17_access(out_dir: Path, created: list[Path]) -> None:
    # q17 : fichier simple pour tester la gestion des erreurs d’accès
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "value"])
    ws.append([1, "ok"])

    save_wb(wb, out_dir / "q17_access.xlsx", created)


def q18_export_in(out_dir: Path, created: list[Path]) -> None:
    # q18 : input pour tester export propre (index=False)
    wb = Workbook()
    ws = wb.active
    ws.append(["policyID", "statecode"])
    ws.append([1, "FL"])
    ws.append([2, "TX"])

    save_wb(wb, out_dir / "q18_export_in.xlsx", created)


def q19_versions(out_dir: Path, created: list[Path]) -> None:
    # q19 : deux versions à comparer
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(["id", "value"])
    ws1.append([1, "A"])
    ws1.append([2, "B"])
    ws1.append([3, "C"])
    save_wb(wb1, out_dir / "q19_v1.xlsx", created)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["id", "value"])
    ws2.append([1, "A"])
    ws2.append([2, "B2"])
    ws2.append([4, "D"])
    save_wb(wb2, out_dir / "q19_v2.xlsx", created)


def q20_pipeline(out_dir: Path, created: list[Path]) -> None:
    # q20 : input avec doublons + valeurs manquantes + colonnes attendues
    wb = Workbook()
    ws = wb.active
    ws.append(["policyID", "statecode", "county"])
    ws.append([1, "FL", "A"])
    ws.append([1, "FL", "A"])     # doublon
    ws.append([None, "TX", "B"])  # policyID manquant
    ws.append([2, None, "C"])     # statecode manquant

    save_wb(wb, out_dir / "q20_pipeline.xlsx", created)


def main() -> None:
    out_dir = ensure_data_dir()
    created: list[Path] = []

    q01_ids(out_dir, created)
    q02_maxrow(out_dir, created)
    q03_big(out_dir, created)
    q04_readonly(out_dir, created)
    q05_dtype(out_dir, created)
    q06_usecols(out_dir, created)
    q07_nrows(out_dir, created)
    q08_dups(out_dir, created)
    q09_emptyrows(out_dir, created)
    q10_schema(out_dir, created)
    q11_dirty_headers(out_dir, created)
    q12_dates(out_dir, created)
    q13_comma(out_dir, created)
    q14_sheets(out_dir, created)      # sert aussi à la jointure (question 15)
    q16_orphans(out_dir, created)
    q17_access(out_dir, created)
    q18_export_in(out_dir, created)
    q19_versions(out_dir, created)
    q20_pipeline(out_dir, created)

    print("Fichiers créés dans le dossier data :")
    for p in created:
        print("-", p)


if __name__ == "__main__":
    main()