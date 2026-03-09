"""Import et export Excel formaté."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COULEUR_HEADER = "1F4E79"
COULEUR_ALERTE = "FFCCCC"
COULEUR_IMPAIR = "F2F2F2"


def exporter_rapport_complet(chemin: str, stock_service) -> None:
    produits = list(stock_service)
    alertes = stock_service.produits_en_alerte()
    stats = stock_service.stats_categories()

    chemin = str(chemin)
    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        pd.DataFrame([p.to_dict() for p in sorted(produits)]).to_excel(writer, "Stock", index=False)
        pd.DataFrame([p.to_dict() for p in alertes]).to_excel(writer, "Alertes", index=False)
        pd.DataFrame([{"categorie": c, **vals} for c, vals in stats.items()]).to_excel(
            writer, "Statistiques", index=False
        )

    wb = load_workbook(chemin)
    _formater_feuille_stock(wb["Stock"], sorted(produits))
    if "Alertes" in wb.sheetnames:
        _formater_feuille_generique(wb["Alertes"])
    _formater_feuille_generique(wb["Statistiques"])
    wb.save(chemin)


def _formater_feuille_generique(ws) -> None:
    header_fill = PatternFill("solid", fgColor=COULEUR_HEADER)
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    _auto_largeur_colonnes(ws)


def _formater_feuille_stock(ws, produits) -> None:
    _formater_feuille_generique(ws)
    alert_fill = PatternFill("solid", fgColor=COULEUR_ALERTE)
    odd_fill = PatternFill("solid", fgColor=COULEUR_IMPAIR)

    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        produit = produits[i] if i < len(produits) else None
        fill = alert_fill if (produit and produit.est_en_alerte()) else (odd_fill if i % 2 else None)
        if fill:
            for cell in row:
                cell.fill = fill


def _auto_largeur_colonnes(ws) -> None:
    for col in ws.columns:
        largeur = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(largeur + 4, 40)


def importer_bon_commande(chemin: str) -> dict:
    path = Path(chemin)
    if not path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {path}")

    xls = pd.ExcelFile(path)
    feuilles_requises = {"produits", "fournisseur"}
    manquantes = feuilles_requises - set(xls.sheet_names)
    if manquantes:
        raise ValueError(f"Feuilles manquantes dans le bon de commande : {sorted(manquantes)}")

    df_produits = pd.read_excel(path, sheet_name="produits")
    df_fournisseur = pd.read_excel(path, sheet_name="fournisseur", nrows=1)
    return {
        "fournisseur": df_fournisseur.iloc[0].to_dict(),
        "lignes": df_produits.to_dict("records"),
    }
